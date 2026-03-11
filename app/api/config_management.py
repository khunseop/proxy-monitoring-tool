from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Dict, Any, List
import json
import io
import re
import os
import tempfile
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    import xlwings as xw
except ImportError:
    xw = None

from app.database.database import get_db
# ... (rest of imports unchanged)

router = APIRouter()

def _apply_header_style(sheet):
# ... (existing _apply_header_style)

def _read_excel_via_xlwings(content: bytes) -> Dict[str, List[List[Any]]]:
    """DRM 등으로 openpyxl이 읽지 못하는 경우 xlwings(Excel 앱)를 통해 데이터를 읽습니다."""
    if not xw:
        raise Exception("xlwings library is not installed.")
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    
    data_map = {}
    app = xw.App(visible=False)
    try:
        wb = app.books.open(tmp_path)
        for sheet in wb.sheets:
            # used_range.value는 시트의 모든 데이터를 2차원 리스트로 반환합니다.
            data_map[sheet.name] = sheet.used_range.value
        wb.close()
    finally:
        app.quit()
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
    
    return data_map

@router.get("/config/export")
# ... (export_full_config_excel body)

@router.post("/config/import")
async def import_full_config_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """펼쳐진 Excel 구조를 다시 복잡한 시스템 JSON 구조로 조립하여 저장합니다. 
    DRM 대응을 위해 openpyxl 실패 시 xlwings를 통한 Fallback을 지원합니다.
    """
    try:
        content = await file.read()
        
        # 1. 시도: openpyxl (빠르고 효율적임)
        excel_data = {}
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                excel_data[sheet_name] = rows
        except Exception as e:
            # 2. 실패 시 xlwings 시도 (Excel 앱 설치 필요)
            try:
                excel_data = _read_excel_via_xlwings(content)
            except Exception as xw_e:
                raise Exception(f"Excel parsing failed (Openpyxl error: {str(e)} / Xlwings fallback error: {str(xw_e)})")

        # 1. Import Groups
        group_map = {}
        if "1.Groups" in excel_data:
            rows = excel_data["1.Groups"]
            for row in rows[1:]: # 헤더 제외
                if not row or len(row) < 1: continue
                name = str(row[0]) if row[0] else None
                if not name: continue
                desc = str(row[1]) if len(row) > 1 and row[1] else None
                existing = db.query(ProxyGroup).filter(ProxyGroup.name == name).first()
                if not existing:
                    existing = ProxyGroup(name=name, description=desc)
                    db.add(existing); db.commit(); db.refresh(existing)
                group_map[name] = existing.id

        # 2. Interface Data Preparation (ProxyInterfaces)
        proxy_if_map = {}
        if "3.ProxyInterfaces" in excel_data:
            rows = excel_data["3.ProxyInterfaces"]
            for row in rows[1:]:
                if not row or len(row) < 4: continue
                host, if_name, in_oid, out_oid = row[0], row[1], row[2], row[3]
                if not host or not if_name: continue
                if host not in proxy_if_map: proxy_if_map[host] = {}
                proxy_if_map[host][if_name] = {"in_oid": str(in_oid or ""), "out_oid": str(out_oid or "")}

        # 3. Import Proxies
        if "2.Proxies" in excel_data:
            rows = excel_data["2.Proxies"]
            for row in rows[1:]:
                if not row or len(row) < 1: continue
                host = str(row[0]) if row[0] else None
                if not host: continue
                
                # OIDs JSON 조립
                p_oids = None
                if host in proxy_if_map:
                    p_oids = json.dumps({"__interface_oids__": proxy_if_map[host]})

                fields = {
                    "host": host, 
                    "username": str(row[1] or "") if len(row) > 1 else "", 
                    "password": encrypt_string(str(row[2])) if len(row) > 2 and row[2] else None,
                    "group_id": group_map.get(str(row[3])) if len(row) > 3 else None, 
                    "traffic_log_path": str(row[4] or "") if len(row) > 4 else "",
                    "is_active": str(row[5]).upper() == "TRUE" if len(row) > 5 else False, 
                    "description": str(row[6] or "") if len(row) > 6 else "",
                    "oids_json": p_oids
                }
                
                existing = db.query(Proxy).filter(Proxy.host == host).first()
                if existing:
                    for k, v in fields.items():
                        if k == "password" and v is None: continue
                        setattr(existing, k, v)
                else:
                    db.add(Proxy(**fields))
            db.commit()

        # 4. System Config Assembly
        res_oids = {}
        res_th = {}
        if "4.SystemResourceOIDs" in excel_data:
            rows = excel_data["4.SystemResourceOIDs"]
            metric_rev_map = {
                "CPU사용률": "cpu", "메모리사용률": "mem", "디스크사용률": "disk",
                "동시접속수(CC)": "cc", "초당접속수(CS)": "cs",
                "HTTP트래픽": "http", "HTTPS트래픽": "https", "HTTP2트래픽": "http2"
            }
            for row in rows[1:]:
                if not row or len(row) < 1: continue
                m_key = metric_rev_map.get(str(row[0]))
                if m_key:
                    res_oids[m_key] = str(row[1] or "") if len(row) > 1 else ""
                    if len(row) > 2 and row[2] is not None:
                        try: res_th[m_key] = float(row[2])
                        except: pass

        common_if_oids = {}
        common_if_th = {}
        common_if_bw = {}
        if "5.CommonInterfaces" in excel_data:
            rows = excel_data["5.CommonInterfaces"]
            for row in rows[1:]:
                if not row or len(row) < 1: continue
                name = str(row[0]) if row[0] else None
                if not name: continue
                common_if_oids[name] = {"in_oid": str(row[1] or ""), "out_oid": str(row[2] or "")} if len(row) > 2 else {"in_oid": "", "out_oid": ""}
                if len(row) > 3 and row[3] is not None:
                    try: common_if_th[name] = float(row[3])
                    except: pass
                if len(row) > 4 and row[4] is not None:
                    try: common_if_bw[name] = float(row[4])
                    except: pass

        # Finalize Resource Config
        res_oids["__thresholds__"] = res_th
        res_oids["__interface_oids__"] = common_if_oids
        res_oids["__interface_thresholds__"] = common_if_th
        res_oids["__interface_bandwidths__"] = common_if_bw
        
        existing_rc = db.query(ResourceConfig).first()
        if not existing_rc: existing_rc = ResourceConfig(); db.add(existing_rc)
        
        # Global Settings
        if "6.GlobalSettings" in excel_data:
            rows = excel_data["6.GlobalSettings"]
            sb_data = {}
            for row in rows[1:]:
                if not row or len(row) < 3: continue
                name, val = row[1], row[2]
                if name == "CommunityString": existing_rc.community = str(val or "public")
                elif name in ["Port", "Timeout", "HostKeyPolicy"]: sb_data[name] = val
            
            if sb_data:
                existing_sb = db.query(SessionBrowserConfig).first()
                if not existing_sb: existing_sb = SessionBrowserConfig(); db.add(existing_sb)
                if "Port" in sb_data and sb_data["Port"]: existing_sb.ssh_port = int(float(sb_data["Port"]))
                if "Timeout" in sb_data and sb_data["Timeout"]: existing_sb.timeout_sec = int(float(sb_data["Timeout"]))
                if "HostKeyPolicy" in sb_data: existing_sb.host_key_policy = str(sb_data["HostKeyPolicy"])

        existing_rc.oids_json = json.dumps(res_oids)
        db.commit()

        return {"status": "success", "message": "Excel configuration imported and assembled successfully"}
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")
