from fastapi import APIRouter, Depends, HTTPException, Query, WebSocket, WebSocketDisconnect
from sqlalchemy.orm import Session
from typing import Dict, Any, List, Tuple, Optional
import time
from datetime import timedelta, datetime, timezone
import asyncio
from app.utils.time import now_kst, KST_TZ
import json
import logging
import warnings
try:
    from cryptography.utils import CryptographyDeprecationWarning
    warnings.filterwarnings("ignore", category=CryptographyDeprecationWarning)
except Exception:
    pass

from app.database.database import get_db
from app.models.proxy import Proxy
from app.models.resource_usage import ResourceUsage as ResourceUsageModel
from app.schemas.resource_usage import (
    ResourceUsage as ResourceUsageSchema,
    CollectRequest,
    CollectResponse,
)
from app.utils.background_collector import background_collector
from pydantic import BaseModel

# Import collection logic from service layer
from app.services.resource_collector import (
    collect_for_proxy,
    get_interface_config_from_db,
    enforce_resource_usage_retention,
    is_system_interface
)


router = APIRouter()
logger = logging.getLogger(__name__)


@router.post("/resource-usage/collect", response_model=CollectResponse)
async def collect_resource_usage(payload: CollectRequest, db: Session = Depends(get_db)):
    if not payload.oids:
        raise HTTPException(status_code=400, detail="oids mapping is required")
    if not payload.proxy_ids or len(payload.proxy_ids) == 0:
        raise HTTPException(status_code=400, detail="proxy_ids is required and cannot be empty")
    if not payload.community:
        raise HTTPException(status_code=400, detail="community is required")

    logger.info(f"[resource_usage] Collect request received proxy_ids={payload.proxy_ids} oids={list(payload.oids.keys())}")
    
    query = db.query(Proxy).filter(Proxy.is_active == True).filter(Proxy.id.in_(payload.proxy_ids))
    proxies: List[Proxy] = query.all()

    if not proxies:
        logger.warning(f"[resource_usage] No active proxies found for ids={payload.proxy_ids}")
        return CollectResponse(requested=0, succeeded=0, failed=0, errors={}, items=[])

    errors: Dict[int, str] = {}
    collected_data: List[dict] = []

    # Get interface_oids from config
    interface_oids, _, _ = get_interface_config_from_db(db)

    # Gather all SNMP collection tasks
    tasks = [collect_for_proxy(p, payload.oids, payload.community, db=db, interface_oids=interface_oids) for p in proxies]
    results = await asyncio.gather(*tasks, return_exceptions=True)
    
    collected_at_ts = now_kst()
    
    for proxy, result in zip(proxies, results):
        try:
            if isinstance(result, Exception):
                errors[proxy.id] = str(result)
                continue
            proxy_id, metrics, err = result
            if err:
                errors[proxy_id] = err
                continue
            
            interface_mbps_data = metrics.get("interface_mbps")
            interface_mbps_json = json.dumps(interface_mbps_data) if interface_mbps_data else None
            
            # Prepare data for bulk insert
            usage_data = {
                "proxy_id": proxy_id,
                "cpu": metrics.get("cpu"),
                "mem": metrics.get("mem"),
                "cc": metrics.get("cc"),
                "cs": metrics.get("cs"),
                "http": metrics.get("http"),
                "https": metrics.get("https"),
                "http2": metrics.get("http2"),
                "blocked": metrics.get("blocked"),
                "disk": metrics.get("disk"),
                "interface_mbps": interface_mbps_json,
                "community": payload.community,
                "oids_raw": json.dumps(payload.oids),
                "collected_at": collected_at_ts,
                "created_at": collected_at_ts,
                "updated_at": collected_at_ts,
            }
            collected_data.append(usage_data)
        except Exception as e:
            errors[proxy.id] = str(e)

    # Bulk insert for better performance
    collected_models: List[ResourceUsageModel] = []
    if collected_data:
        try:
            db.bulk_insert_mappings(ResourceUsageModel, collected_data)
            db.commit()
            # Refresh models to get IDs (for response)
            # Note: bulk_insert_mappings doesn't return IDs, so we query them back
            for data in collected_data:
                model = db.query(ResourceUsageModel).filter(
                    ResourceUsageModel.proxy_id == data["proxy_id"],
                    ResourceUsageModel.collected_at == data["collected_at"]
                ).order_by(ResourceUsageModel.id.desc()).first()
                if model:
                    collected_models.append(model)
        except Exception as e:
            logger.error(f"[resource_usage] Bulk insert failed, falling back to individual inserts: {e}")
            db.rollback()
            # Fallback to individual inserts
            for data in collected_data:
                try:
                    model = ResourceUsageModel(**data)
                    db.add(model)
                    collected_models.append(model)
                except Exception as e2:
                    logger.error(f"[resource_usage] Failed to insert record: {e2}")
            db.commit()

    # Enforce 90-day retention in the background
    asyncio.create_task(asyncio.to_thread(enforce_resource_usage_retention_wrapper))

    logger.info(f"[resource_usage] Collect completed requested={len(proxies)} succeeded={len(collected_data)} failed={len(errors)}")
    if errors:
        logger.warning(f"[resource_usage] Collection errors: {errors}")

    return CollectResponse(
        requested=len(proxies),
        succeeded=len(collected_data),
        failed=len(errors),
        errors=errors,
        items=collected_data,
    )


def enforce_resource_usage_retention_wrapper(days: int = 90) -> None:
    from app.database.database import SessionLocal
    db = SessionLocal()
    try:
        enforce_resource_usage_retention(db, days)
    finally:
        db.close()


@router.get("/resource-usage", response_model=List[ResourceUsageSchema])
async def list_resource_usage(
    db: Session = Depends(get_db),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
):
    rows = (
        db.query(ResourceUsageModel)
        .order_by(ResourceUsageModel.collected_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return rows


@router.get("/resource-usage/latest/{proxy_id}", response_model=ResourceUsageSchema)
async def latest_resource_usage(proxy_id: int, db: Session = Depends(get_db)):
    row = (
        db.query(ResourceUsageModel)
        .filter(ResourceUsageModel.proxy_id == proxy_id)
        .order_by(ResourceUsageModel.collected_at.desc())
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="No resource usage found for proxy")
    return row


class ActiveInterfaceItem(BaseModel):
    index: str
    name: str
    proxy_id: int
    proxy_host: str


@router.get("/resource-usage/active-interfaces", response_model=List[ActiveInterfaceItem])
async def get_active_interfaces(
    proxy_id: Optional[int] = Query(None, description="Filter by proxy ID"),
    group_id: Optional[int] = Query(None, description="Filter by group ID"),
    limit: int = Query(100, ge=1, le=1000, description="Maximum number of interfaces to return"),
    db: Session = Depends(get_db)
):
    """
    Get list of active interfaces from recent collection data.
    """
    cutoff_time = now_kst() - timedelta(hours=24)
    query = db.query(ResourceUsageModel).filter(ResourceUsageModel.collected_at >= cutoff_time)
    
    if proxy_id:
        query = query.filter(ResourceUsageModel.proxy_id == proxy_id)
    elif group_id:
        query = query.join(Proxy).filter(Proxy.group_id == group_id)
    
    recent_records = query.order_by(ResourceUsageModel.collected_at.desc()).limit(limit * 10).all()
    
    interface_map: Dict[Tuple[int, str], Dict[str, Any]] = {}
    for record in recent_records:
        if not record.interface_mbps: continue
        try:
            interface_data = json.loads(record.interface_mbps) if isinstance(record.interface_mbps, str) else record.interface_mbps
            if not isinstance(interface_data, dict): continue
            
            proxy = db.query(Proxy).filter(Proxy.id == record.proxy_id).first()
            proxy_host = proxy.host if proxy else f"proxy_{record.proxy_id}"
            
            for if_index, if_info in interface_data.items():
                if not isinstance(if_info, dict): continue
                if_name = if_info.get("name", f"IF{if_index}")
                total_mbps = (if_info.get("in_mbps", 0) or 0) + (if_info.get("out_mbps", 0) or 0)
                
                if is_system_interface(if_name) or total_mbps <= 0: continue
                
                key = (record.proxy_id, if_index)
                if key not in interface_map:
                    interface_map[key] = {
                        "index": if_index, "name": if_name, "proxy_id": record.proxy_id, "proxy_host": proxy_host
                    }
        except Exception: continue
    
    result = list(interface_map.values())
    result.sort(key=lambda x: (x["proxy_id"], int(x["index"]) if x["index"].isdigit() else 999999))
    return result[:limit]


@router.get("/resource-usage/history", response_model=List[ResourceUsageSchema])
async def get_resource_usage_history(
    db: Session = Depends(get_db),
    proxy_id: Optional[int] = Query(None),
    proxy_ids: Optional[str] = Query(None),
    start_time: Optional[str] = Query(None),
    end_time: Optional[str] = Query(None),
    limit: int = Query(1000, ge=1, le=10000),
    offset: int = Query(0, ge=0),
):
    query = db.query(ResourceUsageModel)
    if proxy_id:
        query = query.filter(ResourceUsageModel.proxy_id == proxy_id)
    elif proxy_ids:
        try:
            ids = [int(x.strip()) for x in proxy_ids.split(',') if x.strip()]
            if ids: query = query.filter(ResourceUsageModel.proxy_id.in_(ids))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid proxy_ids format.")
    
    if start_time:
        try:
            dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
            if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
            query = query.filter(ResourceUsageModel.collected_at >= dt.astimezone(KST_TZ))
        except ValueError: raise HTTPException(status_code=400, detail="Invalid start_time.")
    
    if end_time:
        try:
            dt = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
            if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
            query = query.filter(ResourceUsageModel.collected_at <= dt.astimezone(KST_TZ))
        except ValueError: raise HTTPException(status_code=400, detail="Invalid end_time.")
    
    rows = query.order_by(ResourceUsageModel.collected_at.desc()).offset(offset).limit(limit).all()
    return rows


class ResourceUsageStatsResponse(BaseModel):
    total_count: int
    oldest_record: Optional[str] = None
    newest_record: Optional[str] = None
    retention_days: int = 90
    records_by_proxy: Dict[int, int] = {}


@router.get("/resource-usage/stats", response_model=ResourceUsageStatsResponse)
async def get_resource_usage_stats(
    db: Session = Depends(get_db),
    proxy_id: Optional[int] = Query(None),
    proxy_ids: Optional[str] = Query(None)
):
    query = db.query(ResourceUsageModel)
    if proxy_id:
        query = query.filter(ResourceUsageModel.proxy_id == proxy_id)
    elif proxy_ids:
        try:
            ids = [int(x.strip()) for x in proxy_ids.split(',') if x.strip()]
            if ids: query = query.filter(ResourceUsageModel.proxy_id.in_(ids))
        except ValueError: pass

    total_count = query.count()
    oldest = query.order_by(ResourceUsageModel.collected_at.asc()).first()
    newest = query.order_by(ResourceUsageModel.collected_at.desc()).first()
    
    from sqlalchemy import func
    records_by_proxy = {}
    if not proxy_id:
        proxy_counts = db.query(ResourceUsageModel.proxy_id, func.count(ResourceUsageModel.id)).group_by(ResourceUsageModel.proxy_id).all()
        records_by_proxy = {pid: count for pid, count in proxy_counts}
    
    return ResourceUsageStatsResponse(
        total_count=total_count,
        oldest_record=oldest.collected_at.isoformat() if oldest else None,
        newest_record=newest.collected_at.isoformat() if newest else None,
        retention_days=90,
        records_by_proxy=records_by_proxy
    )


class DeleteResourceUsageRequest(BaseModel):
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    proxy_id: Optional[int] = None
    older_than_days: Optional[int] = None


class DeleteResourceUsageResponse(BaseModel):
    deleted_count: int
    message: str


@router.delete("/resource-usage", response_model=DeleteResourceUsageResponse)
async def delete_resource_usage(request: DeleteResourceUsageRequest, db: Session = Depends(get_db)):
    query = db.query(ResourceUsageModel)
    if request.proxy_id: query = query.filter(ResourceUsageModel.proxy_id == request.proxy_id)
    
    if request.older_than_days:
        cutoff = now_kst() - timedelta(days=request.older_than_days)
        query = query.filter(ResourceUsageModel.collected_at < cutoff)
    else:
        if request.start_time:
            dt = datetime.fromisoformat(request.start_time.replace('Z', '+00:00'))
            query = query.filter(ResourceUsageModel.collected_at >= dt.astimezone(KST_TZ))
        if request.end_time:
            dt = datetime.fromisoformat(request.end_time.replace('Z', '+00:00'))
            query = query.filter(ResourceUsageModel.collected_at <= dt.astimezone(KST_TZ))
    
    deleted_count = query.delete(synchronize_session=False)
    db.commit()
    return DeleteResourceUsageResponse(deleted_count=deleted_count, message=f"{deleted_count}건 삭제되었습니다.")


@router.get("/resource-usage/export")
async def export_resource_usage(
    db: Session = Depends(get_db),
    proxy_id: Optional[int] = Query(None),
    proxy_ids: Optional[str] = Query(None),
    start_time: Optional[str] = Query(None),
    end_time: Optional[str] = Query(None),
    limit: int = Query(10000, ge=1, le=100000)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    
    query = db.query(ResourceUsageModel)
    if proxy_id: query = query.filter(ResourceUsageModel.proxy_id == proxy_id)
    elif proxy_ids:
        try:
            ids = [int(x.strip()) for x in proxy_ids.split(',') if x.strip()]
            if ids: query = query.filter(ResourceUsageModel.proxy_id.in_(ids))
        except ValueError: pass
    
    if start_time:
        dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
        query = query.filter(ResourceUsageModel.collected_at >= dt.astimezone(KST_TZ))
    if end_time:
        dt = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
        query = query.filter(ResourceUsageModel.collected_at <= dt.astimezone(KST_TZ))
    
    rows = query.order_by(ResourceUsageModel.collected_at.desc()).limit(limit).all()
    proxy_map = {p.id: p.host for p in db.query(Proxy).filter(Proxy.id.in_(set(r.proxy_id for r in rows))).all()}
    
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "MainMetrics"
    ws_main.append(['수집 시간', '프록시 ID', '프록시 호스트', 'CPU (%)', 'MEM (%)', 'Disk (%)', 'CC', 'CS', 'HTTP (Mbps)', 'HTTPS (Mbps)', 'HTTP2 (Mbps)'])
    
    ws_if = wb.create_sheet("InterfaceDetails")
    ws_if.append(['수집 시간', '프록시 ID', '프록시 호스트', '인터페이스 명', 'IN (Mbps)', 'OUT (Mbps)'])
    
    for row in rows:
        proxy_host = proxy_map.get(row.proxy_id, f"#{row.proxy_id}")
        ts_str = row.collected_at.strftime('%Y-%m-%d %H:%M:%S')
        ws_main.append([ts_str, row.proxy_id, proxy_host, row.cpu, row.mem, row.disk, row.cc, row.cs, row.http, row.https, row.http2])
        if row.interface_mbps:
            try:
                if_data = json.loads(row.interface_mbps)
                for if_idx, info in if_data.items():
                    ws_if.append([ts_str, row.proxy_id, proxy_host, info.get("name"), info.get("in_mbps", 0), info.get("out_mbps", 0)])
            except: pass
            
    for ws in [ws_main, ws_if]:
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
    
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=resource_usage_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"})


from fastapi.responses import StreamingResponse

class StartBackgroundCollectRequest(BaseModel):
    proxy_ids: List[int]
    community: str
    oids: Dict[str, str]
    interval_sec: int


class BackgroundCollectStatusResponse(BaseModel):
    task_id: str
    status: str
    started_at: Optional[str] = None
    proxy_ids: Optional[List[int]] = None
    interval_sec: Optional[int] = None


@router.post("/resource-usage/background/start", response_model=BackgroundCollectStatusResponse)
async def start_background_collection(payload: StartBackgroundCollectRequest):
    if not payload.oids or not payload.proxy_ids or not payload.community:
        raise HTTPException(status_code=400, detail="Missing required fields")
    if payload.interval_sec < 5:
        raise HTTPException(status_code=400, detail="interval_sec must be at least 5")
    
    task_id = f"ru_{hash(tuple(sorted(payload.proxy_ids)) + (payload.community,) + tuple(sorted(payload.oids.items())))}"
    if background_collector.is_running(task_id):
        status = background_collector.get_status(task_id)
        return BackgroundCollectStatusResponse(task_id=task_id, status="running", **status)
    
    await background_collector.start_collection(task_id=task_id, proxy_ids=payload.proxy_ids, community=payload.community, oids=payload.oids, interval_sec=payload.interval_sec)
    return BackgroundCollectStatusResponse(task_id=task_id, status="started", **background_collector.get_status(task_id))


class StopBackgroundCollectRequest(BaseModel):
    task_id: str


@router.post("/resource-usage/background/stop")
async def stop_background_collection(payload: StopBackgroundCollectRequest):
    await background_collector.stop_collection(payload.task_id)
    return {"status": "stopped", "task_id": payload.task_id}


@router.get("/resource-usage/background/status")
async def get_background_collection_status(task_id: Optional[str] = Query(None)):
    return background_collector.get_status(task_id)


@router.websocket("/ws/resource-usage/status")
async def websocket_collection_status(websocket: WebSocket):
    await websocket.accept()
    await background_collector.register_websocket(websocket)
    try:
        await websocket.send_json({"type": "initial_status", "data": background_collector.get_status()})
        while True:
            if await websocket.receive_text() == "ping": await websocket.send_json({"type": "pong"})
    except WebSocketDisconnect: pass
    finally: await background_collector.unregister_websocket(websocket)
